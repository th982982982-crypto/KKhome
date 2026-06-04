#!/usr/bin/env python3
# Trích biểu mẫu từ các .docx trong "Tổng hợp" -> file DOC (giữ định dạng gốc) + Excel.
# Mỗi "Mẫu số" = 1 file; VB dùng "Phụ lục" (không mã) -> tách theo Phụ lục.
# Ghi manifest scripts/forms-manifest.json để generator nhúng D.forms.
import os, re, json, shutil, unicodedata
from docx import Document
from docx.oxml.ns import qn
import openpyxl
from openpyxl.styles import Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(ROOT, 'Tổng hợp')
OUT  = os.path.join(ROOT, 'Biểu mẫu xuất')

# slug -> {src, short, mode}
#   src  : đường dẫn .docx (tương đối ROOT; nếu chỉ là tên file -> tìm trong "Tổng hợp")
#   short: tên thư mục ngắn trong "Biểu mẫu xuất"
#   mode : 'auto'     = anchor 'Mẫu số <code>' + 'Kèm theo', fallback tách theo 'Phụ lục' (cũ)
#          'template' = anchor là block MỞ ĐẦU bằng 'Mẫu số …' + 'kèm theo' (mẫu nằm cuối VB)
#          'pl'       = bắt buộc tách theo 'Phụ lục'
FORM_DOCS = {
 'tt18' : {'src': 'Thông-tư-18-2026-TT-BTC.docx', 'short': 'TT18',  'mode': 'auto'},
 'tt69' : {'src': 'Thông-tư-69-2025-TT-BTC.docx', 'short': 'TT69',  'mode': 'auto'},
 'tt158': {'src': 'Thông-tư-158-2025-TT-BTC.docx', 'short': 'TT158', 'mode': 'auto'},
 'nd360': {'src': 'Nghị-định-360-2025-NĐ-CP.docx', 'short': 'NĐ360', 'mode': 'auto'},
 'nd181': {'src': 'Nghị-định-181-2025-NĐ-CP.docx', 'short': 'NĐ181', 'mode': 'auto'},
 'nd144': {'src': 'Nghị-định-144-2026-NĐ-CP.docx', 'short': 'NĐ144', 'mode': 'auto'},
 'tt20' : {'src': 'Thông-tư-20-2026-TT-BTC/Thông-tư-20-2026-TT-BTC.docx', 'short': 'TT20', 'mode': 'template'},
 # TT06: mẫu giấy xx/TXNK trong Phụ lục I (bỏ Phụ lục II = chỉ tiêu thông tin điện tử)
 'tt06' : {'src': 'scripts/.doc-cache/Thông-tư-06-2021-TT-BTC.docx', 'short': 'TT06', 'mode': 'mau-start', 'code_filter': r'TXNK', 'title_list': True},
}

P, TBL = qn('w:p'), qn('w:tbl')
MAU = re.compile(r'M[aẫ]u\s*s[ốo]\s*[:\.]?\s*([0-9][0-9A-Za-zĐĐđ\-\./]*)')
PL  = re.compile(r'^\s*(PHỤ LỤC|Phụ lục)\b[^\n]{0,60}')

def btext(el):
    return ''.join(t.text or '' for t in el.iter(qn('w:t')))

def blocks_of(doc):
    return [c for c in doc.element.body.iterchildren() if c.tag in (P, TBL)]

def ascii_fold(s):
    # Supabase Storage key chỉ chấp nhận ASCII → bỏ dấu, Đ/đ -> D/d
    s = s.replace('Đ', 'D').replace('đ', 'd')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s

def norm_code(raw):
    c = ascii_fold(raw.strip().rstrip('.')).replace('/', '-').replace(' ', '')
    c = re.sub(r'[^0-9A-Za-z\-]', '', c)
    return c or 'MAU'

def is_title(t):
    t = t.strip()
    if len(t) < 6 or len(t) > 120: return False
    if re.match(r'(Mẫu số|Phụ lục|\(?Kèm theo|PHỤ LỤC)', t): return False
    letters = [c for c in t if c.isalpha()]
    if not letters: return False
    upper = sum(1 for c in letters if c == c.upper())
    return upper / len(letters) > 0.8

MAU_START = re.compile(r'^\s*M[aẫ]u\s*s[ốo]\s*[:\.]?\s*[0-9]')

def detect_forms(blocks, want='auto', code_filter=None):
    """Trả về (list (code, title, pl, start_idx, end_idx), mode).
    code_filter: regex (chuỗi) áp lên mã thô — chỉ nhận mẫu có mã khớp (dùng cho mode 'mau-start')."""
    anchors = []
    mode = want
    cf = re.compile(code_filter) if code_filter else None
    pl_heads = [i for i, c in enumerate(blocks) if c.tag == P and PL.match(btext(c))]

    if want == 'template':
        # Mẫu nằm ở cuối VB: block MỞ ĐẦU bằng 'Mẫu số …' và có 'kèm theo'
        for i, c in enumerate(blocks):
            t = btext(c)
            m = MAU.search(t)
            if m and MAU_START.match(t) and 'kèm theo' in t.lower():
                anchors.append((i, norm_code(m.group(1)), 'mau'))
        mode = 'template'
    elif want == 'mau-start':
        # Mẫu trong các PHỤ LỤC: block MỞ ĐẦU bằng 'Mẫu số <code>' (không cần 'kèm theo')
        for i, c in enumerate(blocks):
            t = btext(c)
            m = MAU.search(t)
            if m and MAU_START.match(t):
                raw = m.group(1)
                if cf and not cf.search(raw):
                    continue
                anchors.append((i, norm_code(raw), 'mau'))
        mode = 'mau-start'
    elif want == 'pl':
        for i, c in enumerate(blocks):
            if c.tag == P and PL.match(btext(c)):
                anchors.append((i, None, 'pl'))
        mode = 'pl'
    else:  # auto (mặc định, giữ nguyên hành vi cũ)
        for i, c in enumerate(blocks):
            t = btext(c)
            m = MAU.search(t)
            if m and 'Kèm theo' in t:
                anchors.append((i, norm_code(m.group(1)), 'mau'))
        mode = 'mau'
        if not anchors:
            for i, c in enumerate(blocks):
                if c.tag == P and PL.match(btext(c)):
                    anchors.append((i, None, 'pl'))
            mode = 'pl'
    forms = []
    for k, (idx, code, kind) in enumerate(anchors):
        end = anchors[k + 1][0] if k + 1 < len(anchors) else len(blocks)
        # Tránh tràn sang phụ lục khác: cắt tại heading PHỤ LỤC nằm giữa (mode mau-start)
        if mode == 'mau-start':
            for ph in pl_heads:
                if idx < ph < end:
                    end = ph; break
        # title = dòng in hoa đầu tiên trong phạm vi
        title = ''
        for j in range(idx, end):
            if blocks[j].tag == P:
                tt = btext(blocks[j]).strip()
                if is_title(tt):
                    title = re.sub(r'\s+', ' ', tt)[:120]; break
        pl = ''
        if kind == 'pl':
            pl_txt = re.sub(r'\s+', ' ', btext(blocks[idx]).strip())[:60]
            code = norm_code('PL%d' % (k + 1))
            pl = pl_txt
            if not title: title = pl_txt
        forms.append((code, title or code, pl, idx, end))
    return forms, mode

def write_docx(src_path, out_path, start, end):
    shutil.copy(src_path, out_path)
    doc = Document(out_path)
    bl = blocks_of(doc)
    for i, c in enumerate(bl):
        if i < start or i >= end:
            c.getparent().remove(c)
    doc.save(out_path)

def write_xlsx(doc, out_path, start, end, title):
    bl = blocks_of(doc)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = (re.sub(r'[/\\?*\[\]:]', '-', title)[:28] or 'Mẫu')
    ws.column_dimensions['A'].width = 60
    r = 1
    for i in range(start, end):
        el = bl[i]
        if el.tag == P:
            t = btext(el).strip()
            if t:
                ws.cell(row=r, column=1, value=t); r += 1
        else:  # table
            for tr in el.findall(qn('w:tr')):
                cells = tr.findall(qn('w:tc'))
                for ci, tc in enumerate(cells):
                    txt = '\n'.join(p_t for p_t in (btext(p).strip() for p in tc.findall(qn('w:p'))) if p_t)
                    ws.cell(row=r, column=ci + 1, value=txt)
                r += 1
    wb.save(out_path)

MANIFEST_PATH = os.path.join(ROOT, 'scripts/forms-manifest.json')

CODE_ONLY = re.compile(r'^[0-9]{1,3}[0-9A-Za-zĐĐđ\-\./]*$')

def danh_muc_titles(blocks, code_filter=None):
    """Đọc bảng 'Danh mục biểu mẫu' (Tên biểu mẫu | Số hiệu) -> {norm_code: tên chính thức}.
    Ô chỉ chứa mã (vd '01/TXNK') -> tên là đoạn ngay phía trên (không phải số thứ tự)."""
    cf = re.compile(code_filter) if code_filter else None
    out = {}
    for i, c in enumerate(blocks):
        t = btext(c).strip()
        if 2 <= len(t) <= 14 and '/' in t and CODE_ONLY.match(t) and (not cf or cf.search(t)):
            for j in range(i - 1, max(i - 4, -1), -1):
                pt = btext(blocks[j]).strip()
                if pt and not re.fullmatch(r'\d+', pt) and len(pt) > 8:
                    out[norm_code(t)] = re.sub(r'\s+', ' ', pt)[:160]
                    break
    return out

def resolve_src(src):
    # Đường dẫn có '/' -> tương đối ROOT; nếu chỉ tên file -> tìm trong "Tổng hợp"
    p = os.path.join(ROOT, src) if ('/' in src or os.sep in src) else os.path.join(SRC, src)
    return p

def run(only=None):
    print("Trích biểu mẫu..." + (f" (chỉ {only})" if only else ""))
    # Giữ nguyên các slug khác trong manifest (merge thay vì ghi đè)
    manifest = {}
    if os.path.exists(MANIFEST_PATH):
        manifest = json.load(open(MANIFEST_PATH, encoding='utf-8'))
    total = 0
    targets = {only: FORM_DOCS[only]} if only else FORM_DOCS
    for slug, cfg in targets.items():
        src = resolve_src(cfg['src']); short = cfg['short']; want = cfg.get('mode', 'auto')
        doc = Document(src)
        bl = blocks_of(doc)
        forms, mode = detect_forms(bl, want, cfg.get('code_filter'))
        if not forms:
            print(f"  {slug:7} (0 mẫu) — không phát hiện ranh giới"); continue
        tmap = danh_muc_titles(bl, cfg.get('code_filter')) if cfg.get('title_list') else {}
        dword = os.path.join(OUT, f'Biểu mẫu {short}')
        dxls  = os.path.join(OUT, f'Biểu mẫu {short} (Excel)')
        shutil.rmtree(dword, ignore_errors=True); shutil.rmtree(dxls, ignore_errors=True)
        os.makedirs(dword, exist_ok=True); os.makedirs(dxls, exist_ok=True)
        seen = {}; entries = []
        for code, title, pl, s, e in forms:
            if code in seen:
                seen[code] += 1; code = f"{code}-{seen[code]}"
            else:
                seen[code] = 1
            final_title = tmap.get(code, title)
            write_docx(src, os.path.join(dword, code + '.docx'), s, e)
            write_xlsx(doc, os.path.join(dxls, code + '.xlsx'), s, e, final_title)
            entries.append({'code': code, 'title': final_title, 'pl': pl, 'note': ''})
        manifest[slug] = entries
        total += len(entries)
        print(f"  {slug:7} ({mode}) -> {len(entries)} mẫu : {', '.join(x['code'] for x in entries[:6])}{' …' if len(entries)>6 else ''}")
    json.dump(manifest, open(MANIFEST_PATH, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    print(f"Tổng (lần chạy này): {total} mẫu × (DOC+Excel). Manifest: scripts/forms-manifest.json")

if __name__ == '__main__':
    import sys
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    if arg and arg not in FORM_DOCS:
        print(f"Slug không có trong FORM_DOCS: {arg}. Có: {', '.join(FORM_DOCS)}"); sys.exit(1)
    run(arg)
